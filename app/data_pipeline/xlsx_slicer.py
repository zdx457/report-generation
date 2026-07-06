"""将 xlsx 文件按行切片，每个切片包含表头 + 一行内容，保存为 md 文件。

用法示例：
  python xlsx_slicer.py                                      # 默认处理当前目录下 xlsx 文件
  python xlsx_slicer.py --input ../报告模板                   # 指定输入文件夹
  python xlsx_slicer.py --output ./xlsx_slices               # 指定输出文件夹
"""
import argparse
import os

from openpyxl import load_workbook


def xlsx_to_slices(filepath):
    wb = load_workbook(filepath, read_only=True, data_only=True)
    all_rows = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            all_rows.append([cell if cell is not None else "" for cell in row])
    wb.close()

    if not all_rows:
        return [], []

    header = all_rows[0]
    data_rows = all_rows[1:]
    return header, data_rows


def slice_to_md(header, row, sheet_name="Sheet"):
    lines = []
    lines.append(f"## {sheet_name}")
    lines.append("")
    header_line = "| " + " | ".join(str(h) for h in header) + " |"
    separator = "| " + " | ".join("---" for _ in header) + " |"
    data_line = "| " + " | ".join(str(v) for v in row) + " |"
    lines.append(header_line)
    lines.append(separator)
    lines.append(data_line)
    lines.append("")
    return "\n".join(lines)


def process_file(filepath, output_dir, progress_callback=None):
    def _log(msg, level="info"):
        print(msg, flush=True)
        if progress_callback:
            progress_callback({"level": level, "msg": msg})

    basename = os.path.splitext(os.path.basename(filepath))[0]
    _log(f"读取文件: {os.path.basename(filepath)}")
    header, data_rows = xlsx_to_slices(filepath)

    if not header:
        _log(f"跳过空文件: {filepath}", "error")
        return 0

    count = 0
    for i, row in enumerate(data_rows, start=1):
        md_content = slice_to_md(header, row, sheet_name=basename)
        out_name = f"{basename}_row{i}.md"
        out_path = os.path.join(output_dir, out_name)
        with open(out_path, "w", encoding="utf-8") as f:
            f.write(md_content)
        count += 1

    _log(f"生成 {count} 个切片")
    return count


def main():
    parser = argparse.ArgumentParser(description="将 xlsx 按行切片为 md 文件")
    parser.add_argument("--input", type=str, default=".", help="包含 xlsx 文件的输入文件夹")
    parser.add_argument("--output", type=str, default="./xlsx_slices", help="md 文件输出文件夹")
    args = parser.parse_args()

    input_dir = os.path.abspath(args.input)
    output_dir = os.path.abspath(args.output)

    if not os.path.isdir(input_dir):
        print(f"输入文件夹不存在: {input_dir}", file=__import__("sys").stderr)
        return

    os.makedirs(output_dir, exist_ok=True)

    xlsx_files = [f for f in os.listdir(input_dir) if f.endswith(".xlsx") and not f.startswith("~$")]

    if not xlsx_files:
        print(f"在 {input_dir} 中未找到 xlsx 文件")
        return

    total = 0
    for fname in sorted(xlsx_files):
        fpath = os.path.join(input_dir, fname)
        print(f"处理: {fname}")
        n = process_file(fpath, output_dir)
        print(f"  生成 {n} 个切片")
        total += n

    print(f"\n完成！共处理 {len(xlsx_files)} 个文件，生成 {total} 个 md 切片")
    print(f"输出目录: {output_dir}")


if __name__ == "__main__":
    main()