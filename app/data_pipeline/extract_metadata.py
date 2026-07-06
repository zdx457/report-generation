"""从 xlsx 报告模板中提取去重的元数据字段值。

输出 JSON 文件，供查询改写模块使用，确保用户输入的术语
能被标准化为数据库中实际存在的检查类型、部位、检查项目、诊断结论。

用法：
    python extract_metadata.py                          # 默认扫描 report_template/ 目录
    python extract_metadata.py --input ./report_template
    python extract_metadata.py --output metadata.json
"""
import argparse
import json
import os

from openpyxl import load_workbook


def extract_metadata(input_dir, output_path=None, progress_callback=None):
    """从 xlsx 文件中提取去重的元数据。

    Args:
        input_dir: xlsx 文件所在目录
        output_path: 输出 JSON 文件路径，默认为 input_dir/metadata.json
        progress_callback: 进度回调，接收 dict {"level": "info"|"error"|"done", "msg": str}

    Returns:
        dict: 包含 检查类型, 部位, 检查项目, 诊断结论 四个去重列表
    """
    def _log(msg, level="info"):
        print(msg, flush=True)
        if progress_callback:
            progress_callback({"level": level, "msg": msg})

    check_types = set()
    parts = set()
    projects = set()
    diagnoses = set()

    xlsx_files = [
        f for f in os.listdir(input_dir)
        if f.endswith(".xlsx") and not f.startswith("~$")
    ]

    if not xlsx_files:
        _log(f"在 {input_dir} 中未找到 xlsx 文件", "error")
        return None

    for fname in sorted(xlsx_files):
        fpath = os.path.join(input_dir, fname)
        _log(f"处理: {fname}")
        wb = load_workbook(fpath, read_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row in rows:
            if len(row) < 4:
                continue
            if row[0] and str(row[0]).strip():
                check_types.add(str(row[0]).strip())
            if row[1] and str(row[1]).strip():
                parts.add(str(row[1]).strip())
            if row[2] and str(row[2]).strip():
                projects.add(str(row[2]).strip())
            if row[3] and str(row[3]).strip():
                diagnoses.add(str(row[3]).strip())

        wb.close()

    metadata = {
        "检查类型": sorted(check_types),
        "部位": sorted(parts),
        "检查项目": sorted(projects),
        "诊断结论": sorted(diagnoses),
    }

    if output_path is None:
        output_path = os.path.join(input_dir, "metadata.json")

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(metadata, f, ensure_ascii=False, indent=2)

    _log(f"提取完成:")
    _log(f"  检查类型: {len(check_types)} 种 → {sorted(check_types)}")
    _log(f"  部位: {len(parts)} 种 → {sorted(parts)}")
    _log(f"  检查项目: {len(projects)} 种")
    _log(f"  诊断结论: {len(diagnoses)} 种")
    _log(f"输出文件: {output_path}")
    _log("__DONE__", "done")

    return metadata


if __name__ == "__main__":
    default_input = os.path.join(os.path.dirname(os.path.abspath(__file__)), "report_template")
    parser = argparse.ArgumentParser(description="从 xlsx 报告模板中提取去重的元数据字段值")
    parser.add_argument("--input", type=str, default=default_input, help="xlsx 文件所在目录")
    parser.add_argument("--output", type=str, default=None, help="输出 JSON 文件路径")
    args = parser.parse_args()

    extract_metadata(os.path.abspath(args.input), args.output)