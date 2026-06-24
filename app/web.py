"""RAG 问答 Web 界面（Gradio）。

用法示例：
  python web.py
  python web.py --share          # 生成公网链接
  python web.py --top-k 5        # 调整检索数量
  python web.py --debug          # 显示检索详情
"""
import json
import os
import shutil
import sys
import time

import gradio as gr
import requests
from dotenv import load_dotenv
from openpyxl import load_workbook
from pymilvus import CollectionSchema, DataType, FieldSchema, MilvusClient

from rerank import get_rerank_config, rerank_documents
from retrieval import (
    multi_recall,
)
from query_rewrite import rewrite_query, needs_rewrite, is_too_vague, get_clarification, standardize_query, parse_query_keywords

load_dotenv(os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", ".env"))

EMBED_URL = os.environ.get("EMBED_URL", "http://14.22.83.225:11002/v1/embeddings")
EMBED_MODEL = os.environ.get("EMBED_MODEL", "bge-m3")
CHAT_URL = os.environ.get("CHAT_URL", "http://14.22.86.97:11001/v1/chat/completions")
CHAT_MODEL = os.environ.get("CHAT_MODEL", "qwen36_27b_lora")
DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "milvus_lite.db")
COLLECTION_NAME = "report_slices"
REPORT_TEMPLATE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "report_template")
XLSX_SLICES_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "xlsx_slices")
EMBED_DIM = 1024

PROMPT_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "prompt.md")


def load_system_prompt():
    if os.path.exists(PROMPT_FILE):
        with open(PROMPT_FILE, "r", encoding="utf-8") as f:
            return f.read().strip()
    return "你是一个医疗影像报告生成助手。请根据检索到的参考信息回答用户问题。如果参考信息不足以回答问题，请如实说明。"


SYSTEM_PROMPT = load_system_prompt()

milvus_client = MilvusClient(uri=DB_PATH)
milvus_client.load_collection(COLLECTION_NAME)


def get_embedding(text):
    payload = {"model": EMBED_MODEL, "input": [text]}
    r = requests.post(EMBED_URL, json=payload, timeout=30)
    r.raise_for_status()
    return r.json()["data"][0]["embedding"]


def search(query_vector, top_k=3):
    results = milvus_client.search(
        collection_name=COLLECTION_NAME,
        data=[query_vector],
        limit=top_k,
        output_fields=["text", "source", "检查类型", "部位", "检查项目", "诊断结论"],
    )
    return results[0]


def _xlsx_to_slices(filepath):
    wb = load_workbook(filepath, read_only=True, data_only=True)
    all_rows = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            all_rows.append([cell if cell is not None else "" for cell in row])
    wb.close()
    if not all_rows:
        return [], []
    header = all_rows[0]
    data_rows = all_rows[1:]
    return header, data_rows


def _slice_to_md(header, row, sheet_name="Sheet"):
    lines = [f"## {sheet_name}", ""]
    header_line = "| " + " | ".join(str(h) for h in header) + " |"
    separator = "| " + " | ".join("---" for _ in header) + " |"
    data_line = "| " + " | ".join(str(v) for v in row) + " |"
    lines.extend([header_line, separator, data_line, ""])
    return "\n".join(lines)


def _parse_md_slice(filepath):
    with open(filepath, "r", encoding="utf-8") as f:
        content = f.read()
    lines = [l.strip() for l in content.strip().split("\n") if l.strip()]
    header = None
    data_row = None
    for line in lines:
        if line.startswith("|") and "---" in line:
            continue
        if line.startswith("##"):
            continue
        cells = [c.strip() for c in line.split("|")]
        cells = [c for c in cells if c]
        if not cells:
            continue
        if header is None:
            header = cells
        else:
            data_row = cells
            break
    if not header or not data_row:
        return None, None
    text_parts = []
    metadata = {}
    for h, v in zip(header, data_row):
        text_parts.append(f"{h}：{v}")
        if h in ("检查类型", "部位", "检查项目", "诊断结论"):
            metadata[h] = v
    return "\n".join(text_parts), metadata


def _get_embeddings_batch(texts, batch_size=16):
    all_embeddings = []
    total = len(texts)
    for i in range(0, total, batch_size):
        batch = texts[i:i + batch_size]
        payload = {"model": EMBED_MODEL, "input": batch}
        for attempt in range(5):
            try:
                r = requests.post(EMBED_URL, json=payload, timeout=120)
                r.raise_for_status()
                data = r.json()["data"]
                data.sort(key=lambda x: x["index"])
                all_embeddings.extend([d["embedding"] for d in data])
                break
            except Exception as e:
                if attempt < 4:
                    time.sleep((attempt + 1) * 5)
                else:
                    raise RuntimeError(f"向量化批次失败: {e}") from e
        time.sleep(0.2)
    return all_embeddings


def _create_collection_if_needed(client):
    if client.has_collection(COLLECTION_NAME):
        return
    schema = CollectionSchema(fields=[
        FieldSchema(name="id", dtype=DataType.INT64, is_primary=True, auto_id=True),
        FieldSchema(name="vector", dtype=DataType.FLOAT_VECTOR, dim=EMBED_DIM),
        FieldSchema(name="text", dtype=DataType.VARCHAR, max_length=4096),
        FieldSchema(name="source", dtype=DataType.VARCHAR, max_length=512),
        FieldSchema(name="检查类型", dtype=DataType.VARCHAR, max_length=256),
        FieldSchema(name="部位", dtype=DataType.VARCHAR, max_length=256),
        FieldSchema(name="检查项目", dtype=DataType.VARCHAR, max_length=256),
        FieldSchema(name="诊断结论", dtype=DataType.VARCHAR, max_length=1024),
    ])
    index_params = client.prepare_index_params()
    index_params.add_index(field_name="vector", index_type="IVF_FLAT", metric_type="COSINE", params={"nlist": 128})
    client.create_collection(collection_name=COLLECTION_NAME, schema=schema, index_params=index_params)


def upload_and_process(files, progress=gr.Progress()):
    if not files:
        return "请上传 xlsx 文件"

    os.makedirs(REPORT_TEMPLATE_DIR, exist_ok=True)
    os.makedirs(XLSX_SLICES_DIR, exist_ok=True)

    saved_files = []
    for f in files:
        fname = os.path.basename(f.name)
        if not fname.endswith(".xlsx") or fname.startswith("~$"):
            continue
        dest = os.path.join(REPORT_TEMPLATE_DIR, fname)
        shutil.copy2(f.name, dest)
        saved_files.append((fname, dest))

    if not saved_files:
        return "未找到有效的 xlsx 文件（跳过临时文件和非 xlsx 文件）"

    progress(0.1, desc="切片中...")
    total_slices = 0
    new_md_files = []

    for idx, (fname, fpath) in enumerate(saved_files):
        progress((0.1 + 0.2 * idx / len(saved_files)), desc=f"切片: {fname}")
        header, data_rows = _xlsx_to_slices(fpath)
        if not header:
            continue
        basename = os.path.splitext(fname)[0]
        for i, row in enumerate(data_rows, start=1):
            md_content = _slice_to_md(header, row, sheet_name=basename)
            out_name = f"{basename}_row{i}.md"
            out_path = os.path.join(XLSX_SLICES_DIR, out_name)
            with open(out_path, "w", encoding="utf-8") as f:
                f.write(md_content)
            new_md_files.append(out_name)
            total_slices += 1

    if total_slices == 0:
        return "切片结果为空，请检查 xlsx 文件内容"

    progress(0.3, desc="检查增量...")
    client = MilvusClient(uri=DB_PATH)
    _create_collection_if_needed(client)

    existing = set()
    try:
        client.load_collection(COLLECTION_NAME)
        results = client.query(COLLECTION_NAME, filter="", output_fields=["source"])
        existing = {r["source"] for r in results}
    except Exception:
        pass

    truly_new = [f for f in new_md_files if f not in existing]
    if not truly_new:
        client.close()
        return f"切片完成: {total_slices} 个，但全部已存在于数据库中，无需更新"

    progress(0.4, desc=f"解析 {len(truly_new)} 个新切片...")
    texts = []
    metadatas = []
    source_files = []
    for fname in truly_new:
        fpath = os.path.join(XLSX_SLICES_DIR, fname)
        text, meta = _parse_md_slice(fpath)
        if text:
            texts.append(text)
            metadatas.append(meta or {})
            source_files.append(fname)

    if not texts:
        client.close()
        return "解析完成但无有效切片"

    progress(0.5, desc=f"向量化 {len(texts)} 条切片...")
    embeddings = _get_embeddings_batch(texts, batch_size=16)

    progress(0.85, desc="写入数据库...")
    data = []
    for i in range(len(texts)):
        meta = metadatas[i]
        data.append({
            "vector": embeddings[i],
            "text": texts[i],
            "source": source_files[i],
            "检查类型": meta.get("检查类型", ""),
            "部位": meta.get("部位", ""),
            "检查项目": meta.get("检查项目", ""),
            "诊断结论": meta.get("诊断结论", ""),
        })

    insert_batch = 500
    for i in range(0, len(data), insert_batch):
        client.insert(collection_name=COLLECTION_NAME, data=data[i:i + insert_batch])
    client.close()

    global milvus_client
    milvus_client = MilvusClient(uri=DB_PATH)
    milvus_client.load_collection(COLLECTION_NAME)

    progress(1.0, desc="完成")
    return (
        f"处理完成！\n"
        f"- 上传文件: {len(saved_files)} 个\n"
        f"- 生成切片: {total_slices} 个\n"
        f"- 新增入库: {len(truly_new)} 条\n"
        f"- 已有跳过: {len(new_md_files) - len(truly_new)} 条"
    )


def chat_stream(messages, max_tokens=1024, temperature=0.7):
    payload = {
        "model": CHAT_MODEL,
        "messages": messages,
        "max_tokens": max_tokens,
        "temperature": temperature,
        "stream": True,
    }
    headers = {"Content-Type": "application/json"}

    full_reply = ""
    with requests.post(CHAT_URL, headers=headers, json=payload, stream=True) as r:
        r.raise_for_status()
        for line in r.iter_lines(decode_unicode=True):
            if not line:
                continue
            if line.startswith("data:"):
                data = line[len("data:"):].strip()
            else:
                data = line
            if data == "[DONE]":
                break
            try:
                obj = json.loads(data)
                if isinstance(obj, dict) and "choices" in obj:
                    for c in obj["choices"]:
                        delta = c.get("delta", {})
                        content = delta.get("content", "")
                        if content:
                            full_reply += content
                            yield full_reply
            except Exception:
                pass


def _esc(s):
    return s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


class StepCounter:
    def __init__(self):
        self.n = 0

    def next(self, title, status=""):
        self.n += 1
        status_str = f" {status}" if status else ""
        return f"<p><b>{self.n}. {title}</b>{status_str}</p>\n"


def rag_respond(message, history, top_k, rerank_top_k, temperature):
    if rerank_top_k > top_k:
        rerank_top_k = top_k

    step = StepCounter()

    if is_too_vague(message):
        clarification = get_clarification(message)
        thinking_html = "<details open><summary>⚠️ 查询过于模糊</summary>\n"
        thinking_html += "<div style='font-size:14px;'>\n"
        thinking_html += "<p><b>您的输入过于模糊，请补充检查部位或诊断信息：</b></p>\n"
        for line in clarification.split("\n"):
            thinking_html += f"<div style='margin-left:16px;'>{line}</div>\n"
        thinking_html += "</div>\n</details>\n"
        yield thinking_html
        return

    thinking_html = "<details open><summary>🧠 思考过程</summary>\n"
    thinking_html += "<div style='font-size:15px;'>\n"

    # ── 1. 用户输入 ──
    thinking_html += step.next("用户输入")
    thinking_html += f"<div style='margin-left:16px;font-size:13px;color:#666;'><code>{_esc(message)}</code></div>\n"

    # ── 2. 查询改写 ──
    original_query = message
    standardized_query = standardize_query(message)
    rewritten_query = standardized_query
    query_was_rewritten = False

    if needs_rewrite(standardized_query):
        thinking_html += step.next("查询改写", "⏳ 正在改写...")
        yield thinking_html + "</div>\n</details>\n"

        rewritten_query = rewrite_query(standardized_query)
        if rewritten_query != standardized_query:
            query_was_rewritten = True
            thinking_html = thinking_html.replace(" ⏳ 正在改写...", "")
            if standardized_query != message:
                thinking_html += "<div style='margin-left:16px;font-size:13px;color:#666;'>原始查询: <code>{}</code> → 标准化: <code>{}</code> → 改写为: <code>{}</code></div>\n".format(
                    _esc(original_query), _esc(standardized_query), _esc(rewritten_query),
                )
            else:
                thinking_html += "<div style='margin-left:16px;font-size:13px;color:#666;'>原始查询: <code>{}</code> → 改写为: <code>{}</code></div>\n".format(
                    _esc(original_query), _esc(rewritten_query),
                )
        else:
            thinking_html = thinking_html.replace(" ⏳ 正在改写...", "")
            thinking_html += "<div style='margin-left:16px;font-size:13px;color:#666;'>改写结果与标准化查询相同，无需改写</div>\n"
    else:
        thinking_html += step.next("查询改写")
        if standardized_query != message:
            thinking_html += "<div style='margin-left:16px;font-size:13px;color:#666;'>原始查询: <code>{}</code> → 标准化: <code>{}</code>，无需改写</div>\n".format(
                _esc(original_query), _esc(standardized_query),
            )
        else:
            thinking_html += "<div style='margin-left:16px;font-size:13px;color:#666;'>查询足够具体，无需改写</div>\n"

    search_query = rewritten_query if query_was_rewritten else standardized_query

    # ── 3. 多路召回 ──
    thinking_html += step.next("多路召回", "⏳ 正在向量化...")
    yield thinking_html + "</div>\n</details>\n"

    query_vec = get_embedding(search_query)

    thinking_html = thinking_html.replace(" ⏳ 正在向量化...", "")
    thinking_html += "<div style='margin-left:16px;font-size:13px;color:#666;'>向量化模型: <code>{}</code> | 维度: {}</div>\n".format(EMBED_MODEL, len(query_vec))

    keywords = parse_query_keywords(standardized_query)
    thinking_html += "<div style='margin-left:16px;font-size:13px;color:#666;'>解析关键词: 检查类型=<code>{}</code> | 部位=<code>{}</code> | 诊断关键词=<code>{}</code></div>\n".format(
        keywords.get("检查类型", "-") or "-",
        keywords.get("部位", "-") or "-",
        ", ".join(keywords.get("诊断关键词", [])) or "-",
    )

    # ── 路径1: 向量检索 ──
    thinking_html += "<div style='margin-left:16px;margin-top:8px;'><b>📌 路径1: 向量检索</b>（语义相似度匹配）⏳ 正在检索...</div>\n"
    yield thinking_html + "</div>\n</details>\n"

    candidates_list, recall_details = multi_recall(
        query_vec, keywords, top_k, milvus_client, return_details=True,
    )

    vec_results = recall_details.get("vector", [])
    thinking_html = thinking_html.replace("⏳ 正在检索...", f"✅ 返回 {len(vec_results)} 条")
    for i, c in enumerate(vec_results[:5], 1):
        dist_str = f"相似度: {c.get('_distance', -1):.4f}"
        thinking_html += (
            "<div style='margin-left:32px;margin-bottom:1px;font-size:13px;color:#555;'>"
            "候选{i} · {dist} · 来源: <code>{source}</code></div>\n"
        ).format(i=i, dist=dist_str, source=_esc(c["source"]))

    # ── 路径2: 元数据过滤 ──
    thinking_html += "<div style='margin-left:16px;margin-top:8px;'><b>📌 路径2: 元数据过滤</b>（检查类型/部位/诊断精确匹配）⏳ 正在检索...</div>\n"
    yield thinking_html + "</div>\n</details>\n"

    meta_results = recall_details.get("metadata", [])

    filter_desc_parts = []
    if keywords.get("检查类型"):
        filter_desc_parts.append(f'检查类型=="{keywords["检查类型"]}"')
    if keywords.get("部位"):
        filter_desc_parts.append(f'部位=="{keywords["部位"]}"')
    if keywords.get("诊断关键词"):
        for kw in keywords["诊断关键词"]:
            filter_desc_parts.append(f'诊断结论 LIKE "%{kw}%"')
    filter_desc = " AND ".join(filter_desc_parts) if filter_desc_parts else "无条件"

    thinking_html = thinking_html.replace("⏳ 正在检索...", f"✅ 返回 {len(meta_results)} 条")
    thinking_html += f"<div style='margin-left:32px;font-size:13px;color:#888;'>过滤条件: <code>{_esc(filter_desc)}</code></div>\n"
    for i, c in enumerate(meta_results[:5], 1):
        thinking_html += (
            "<div style='margin-left:32px;margin-bottom:1px;font-size:13px;color:#555;'>"
            "候选{i} · 精确匹配 · 来源: <code>{source}</code></div>\n"
        ).format(i=i, source=_esc(c["source"]))

    # ── 路径3: 关键词检索 ──
    thinking_html += "<div style='margin-left:16px;margin-top:8px;'><b>📌 路径3: 关键词检索</b>（全文 like 匹配）⏳ 正在检索...</div>\n"
    yield thinking_html + "</div>\n</details>\n"

    kw_results = recall_details.get("keyword", [])

    kw_desc_parts = []
    if keywords.get("诊断关键词"):
        for kw in keywords["诊断关键词"]:
            kw_desc_parts.append(f'text LIKE "%{kw}%"')
    if keywords.get("检查类型"):
        ct = keywords["检查类型"]
        if ct == "MRI":
            ct = "MR"
        kw_desc_parts.append(f'检查类型=="{ct}"')
    if keywords.get("部位"):
        kw_desc_parts.append(f'部位=="{keywords["部位"]}"')
    kw_desc = " AND ".join(kw_desc_parts) if kw_desc_parts else "无条件"

    thinking_html = thinking_html.replace("⏳ 正在检索...", f"✅ 返回 {len(kw_results)} 条")
    thinking_html += f"<div style='margin-left:32px;font-size:13px;color:#888;'>搜索条件: <code>{_esc(kw_desc)}</code></div>\n"
    for i, c in enumerate(kw_results[:5], 1):
        thinking_html += (
            "<div style='margin-left:32px;margin-bottom:1px;font-size:13px;color:#555;'>"
            "候选{i} · 全文匹配 · 来源: <code>{source}</code></div>\n"
        ).format(i=i, source=_esc(c["source"]))

    # ── 合并去重 ──
    total_before = len(vec_results) + len(meta_results) + len(kw_results)
    thinking_html += "<div style='margin-left:16px;margin-top:8px;font-size:13px;color:#333;'><b>📊 合并去重</b>: {} 条 → {} 条（去重 {} 条）</div>\n".format(
        total_before, len(candidates_list), total_before - len(candidates_list),
    )

    # ── 4. Rerank 重排序 ──
    thinking_html += step.next("Rerank 重排序", "⏳ 正在精排...")
    yield thinking_html + "</div>\n</details>\n"

    documents = [e["text"] for e in candidates_list]

    reranked_entities = []
    try:
        rerank_results = rerank_documents(message, documents, top_n=rerank_top_k)
        for rr in rerank_results:
            idx = rr.get("index", 0)
            rerank_score = rr.get("relevance_score", 0)
            entity = candidates_list[idx]
            entity["_rerank_score"] = rerank_score
            reranked_entities.append(entity)
    except Exception as e:
        reranked_entities = candidates_list[:rerank_top_k]
        for entity in reranked_entities:
            entity["_rerank_score"] = -1

    ref_details = []
    contexts = []
    for i, entity in enumerate(reranked_entities, 1):
        vec_score = entity.get("_distance", 0)
        rerank_score = entity.get("_rerank_score", 0)
        recall_path = entity.get("_recall_path", "vector")
        ref_details.append({
            "index": i,
            "source": entity["source"],
            "vec_score": f"{vec_score:.4f}",
            "rerank_score": f"{rerank_score:.4f}",
            "recall_path": recall_path,
            "检查类型": entity.get("检查类型", ""),
            "部位": entity.get("部位", ""),
            "检查项目": entity.get("检查项目", ""),
            "诊断结论": entity.get("诊断结论", ""),
            "text": entity["text"],
        })
        contexts.append(f"【参考{i}】(Rerank相关性分数: {rerank_score:.4f}，来源: {entity['source']})\n{entity['text']}")

    context_text = "\n\n".join(contexts)

    thinking_html = thinking_html.replace(" ⏳ 正在精排...", "")
    thinking_html += "<div style='margin-left:16px;font-size:13px;color:#666;'>模型: <code>{}</code> | top-{} 命中</div>\n".format(get_rerank_config()["rerank_model"], rerank_top_k)
    rerank_failed = any(ref["rerank_score"] == "-1.0000" for ref in ref_details)
    if rerank_failed:
        thinking_html += "<div style='margin-left:16px;font-size:13px;color:#e74c3c;'>⚠️ Rerank 调用失败，已降级为向量检索结果</div>\n"
    for ref in ref_details:
        thinking_html += (
            "<div style='margin-left:16px;margin-bottom:8px;'>"
            "<b>参考{index}</b> · 召回路径: {recall_path} · Rerank分数: {rerank_score} · 来源: <code>{source}</code><br>"
            "检查类型: {检查类型} | 部位: {部位} | 检查项目: {检查项目}<br>"
            "诊断结论: {诊断结论}<br>"
            "<details><summary>查看全文</summary><pre style='white-space:pre-wrap;'>{text}</pre></details>"
            "</div>\n"
        ).format(**ref)

    # ── 5. 拼接提示词 ──
    user_message = f"参考信息：\n{context_text}\n\n用户问题：{message}"
    messages = [
        {"role": "system", "content": SYSTEM_PROMPT},
        {"role": "user", "content": user_message},
    ]

    thinking_html += step.next("拼接提示词")
    thinking_html += "<div style='margin-left:16px;margin-bottom:8px;'>\n"
    thinking_html += "<details><summary>System Prompt</summary><pre style='white-space:pre-wrap;'>{}</pre></details>\n".format(
        _esc(SYSTEM_PROMPT)
    )
    thinking_html += "<details><summary>User Prompt</summary><pre style='white-space:pre-wrap;'>{}</pre></details>\n".format(
        _esc(user_message)
    )
    thinking_html += "</div>\n"

    # ── 6. 最终回复 ──
    thinking_html += step.next("最终回复")
    thinking_html += "<div style='margin-left:16px;margin-bottom:8px;'>"
    thinking_html += "模型: <code>{}</code> | 温度: {}</div>\n".format(CHAT_MODEL, temperature)

    thinking_html += "</div>\n</details>\n"

    yield thinking_html

    collapsed_thinking = thinking_html.replace("<details open>", "<details>")

    partial_reply = ""
    for chunk in chat_stream(messages, temperature=temperature):
        partial_reply = chunk
        yield collapsed_thinking + "\n" + partial_reply


def build_ui():
    with gr.Blocks(title="医疗影像报告生成", css="""
        .chatbot .message { font-size: 16px; line-height: 1.6; }

        details {
            border: 1px solid #e0e0e0;
            border-radius: 8px;
            background: #fafafa;
            margin: 8px 0;
            overflow: hidden;
        }
        details[open] {
            border-color: #4a90d9;
            box-shadow: 0 2px 8px rgba(74,144,217,0.12);
        }
        details > summary {
            padding: 10px 14px;
            font-size: 14px;
            font-weight: 600;
            cursor: pointer;
            user-select: none;
            list-style: none;
            display: flex;
            align-items: center;
            gap: 8px;
            transition: background 0.2s;
        }
        details > summary:hover {
            background: #f0f4f8;
        }
        details > summary::-webkit-details-marker {
            display: none;
        }
        details > summary::before {
            content: '▶';
            font-size: 10px;
            color: #4a90d9;
            transition: transform 0.25s ease;
            flex-shrink: 0;
        }
        details[open] > summary::before {
            transform: rotate(90deg);
        }
        details > div {
            padding: 0 14px 10px 14px;
        }
    """) as demo:
        gr.Markdown("# 🏥 医疗影像报告生成系统")
        gr.Markdown(
            f"Embedding: `{EMBED_MODEL}` | Rerank: `{get_rerank_config()['rerank_model']}` | 生成模型: `{CHAT_MODEL}`"
        )

        with gr.Row():
            with gr.Column(scale=3):
                chatbot = gr.Chatbot(height=800, sanitize_html=False)
                with gr.Row():
                    msg_input = gr.Textbox(
                        placeholder="输入问题，如：CT弥漫性肺气肿",
                        show_label=False,
                        scale=4,
                    )
                    send_btn = gr.Button("发送", variant="primary", scale=1)
                    clear_btn = gr.Button("清空", scale=1)

            with gr.Column(scale=1):
                gr.Markdown("### 参数设置")
                top_k = gr.Slider(1, 20, value=5, step=1, label="向量检索数量 (Top-K)")
                rerank_top_k = gr.Slider(1, 20, value=3, step=1, label="Rerank 返回数量 (Top-K)", info="Rerank精排后返回给LLM的候选数量")
                temperature = gr.Slider(0.1, 1.0, value=0.7, step=0.1, label="温度 (Temperature)")

                gr.Markdown("---")
                gr.Markdown("### 📤 上传报告模板")
                file_upload = gr.File(
                    label="上传 xlsx 文件",
                    file_count="multiple",
                    file_types=[".xlsx"],
                )
                upload_btn = gr.Button("上传并处理", variant="secondary")
                upload_status = gr.Textbox(label="处理状态", interactive=False, lines=5)

        def respond(message, history, top_k_val, rerank_top_k_val, temp_val):
            display = history + [
                {"role": "user", "content": message},
            ]
            yield display, ""

            reply_text = ""
            for partial in rag_respond(message, history, top_k_val, rerank_top_k_val, temp_val):
                reply_text = partial
                display = history + [
                    {"role": "user", "content": message},
                    {"role": "assistant", "content": reply_text},
                ]
                yield display, ""

        def clear_chat():
            return []

        msg_input.submit(
            respond,
            inputs=[msg_input, chatbot, top_k, rerank_top_k, temperature],
            outputs=[chatbot, msg_input],
        )

        send_btn.click(
            respond,
            inputs=[msg_input, chatbot, top_k, rerank_top_k, temperature],
            outputs=[chatbot, msg_input],
        )

        clear_btn.click(clear_chat, outputs=[chatbot])

        upload_btn.click(
            upload_and_process,
            inputs=[file_upload],
            outputs=[upload_status],
        )

    return demo


def main():
    share = "--share" in sys.argv
    debug = "--debug" in sys.argv

    demo = build_ui()
    server_name = os.environ.get("GRADIO_SERVER_NAME", "127.0.0.1")
    server_port = int(os.environ.get("GRADIO_SERVER_PORT", "7860"))
    demo.launch(share=share, debug=debug, theme=gr.themes.Soft(), server_name=server_name, server_port=server_port)


if __name__ == "__main__":
    main()